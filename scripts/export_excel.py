#!/usr/bin/env python3
"""Chạy tất cả tool với sample và ghi Excel kết quả."""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.run_tool import SAMPLES  # noqa: E402
from tools.registry import REGISTRY, get_tool  # noqa: E402


def _flatten(obj, prefix=""):
    rows = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, (dict, list)):
                rows.extend(_flatten(v, key))
            else:
                rows.append((key, v))
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            key = f"{prefix}[{i}]"
            if isinstance(item, (dict, list)):
                rows.extend(_flatten(item, key))
            else:
                rows.append((key, item))
    else:
        rows.append((prefix, obj))
    return rows


def build_workbook(results: dict[str, dict]) -> Workbook:
    wb = Workbook()
    # Index
    ws = wb.active
    ws.title = "Tong_quan"
    header = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws.append(["STT", "Tool ID", "Tên công cụ", "Danh mục", "Trạng thái chạy", "Thời điểm"])
    for cell in ws[1]:
        cell.fill = header
        cell.font = header_font
    now = datetime.now().isoformat(timespec="seconds")
    for i, (tool_id, meta) in enumerate(REGISTRY.items(), 1):
        ok = "OK" if tool_id in results and "error" not in results[tool_id] else "LỖI"
        ws.append([i, tool_id, meta["title"], meta["category"], ok, now])

    # Catalog sheet
    cat = wb.create_sheet("Danh_muc_thiet_ke")
    cat.append(["Tool ID", "Tên", "Danh mục", "Mô tả ngắn", "Input chính", "Output chính"])
    for cell in cat[1]:
        cell.fill = header
        cell.font = header_font
    designs = {
        "luong_gross_net": ("Gross/Net/Phụ thuộc/Vùng", "Net, thuế, BH"),
        "quyet_toan_tncn": ("Thu nhập năm, BH, đã khấu trừ", "Nộp thêm/Hoàn"),
        "bao_hiem_that_nghiep": ("Lương BQ, tháng đóng", "Trợ cấp ước tính"),
        "bao_hiem_huu_tri": ("Lương BQ, năm đóng, quỹ TV", "Lương hưu + rút 4%"),
        "gia_vang_loi_lo": ("SL, giá mua/bán", "Lãi/lỗ"),
        "dca": ("Tiền/tháng, số tháng, lợi suất", "FV DCA vs Lump Sum"),
        "quy_du_phong": ("Chi tiêu, mục tiêu tháng", "Thiếu / số tháng"),
        "ty_gia": ("Số tiền, tỷ giá", "VND hoặc ngoại tệ"),
        "so_sanh_dau_tu_vs_tiet_kiem": ("Vốn, năm, lãi", "FV 2 kênh"),
        "tiet_kiem_vs_lam_phat": ("Vốn, lãi, lạm phát", "Sức mua thực"),
        "diem_suc_khoe": ("Thu nhập, chi, nợ, đệm", "Score 0-100"),
        "rui_ro_tai_chinh": ("Thu nhập, đệm, rủi ro TS", "Risk score"),
        "fire": ("Chi tiêu năm, tiết kiệm", "FIRE number, số năm"),
        "barista_fire": ("Chi tiêu, part-time", "Barista number"),
        "mua_vs_thue_nha": ("Giá nhà, thuê, vay", "So sánh NW"),
        "ty_le_no_dti": ("Thu nhập, trả nợ", "DTI %"),
        "snowball_avalanche": ("Danh sách nợ", "Tháng & lãi"),
        "vay_mua_xe": ("Giá xe, vay, OPEX", "PMT + tổng chi phí"),
        "so_sanh_khoan_vay": ("Nhiều khoản vay", "Tổng trả tốt nhất"),
        "lai_the_tin_dung": ("Dư nợ, lãi năm", "Tổng lãi"),
        "appreciation_bds": ("Giá mua, CAGR", "Giá trị tương lai"),
        "chi_phi_nuoi_con": ("Chi phí tháng, tuổi", "Tổng 0-18"),
    }
    for tool_id, meta in REGISTRY.items():
        inp, out = designs.get(tool_id, ("xem module", "dict JSON"))
        cat.append([tool_id, meta["title"], meta["category"], meta["title"], inp, out])

    # One sheet per tool (flattened)
    for tool_id, result in results.items():
        name = tool_id[:28]
        sheet = wb.create_sheet(name)
        sheet.append(["Key", "Value"])
        for cell in sheet[1]:
            cell.fill = header
            cell.font = header_font
        for k, v in _flatten(result):
            sheet.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 48

    # Raw JSON sheet
    raw = wb.create_sheet("Raw_JSON")
    raw.append(["tool_id", "json"])
    for tool_id, result in results.items():
        raw.append([tool_id, json.dumps(result, ensure_ascii=False)])

    for sheet in wb.worksheets:
        for col in range(1, min(sheet.max_column, 10) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = max(
                14, sheet.column_dimensions[get_column_letter(col)].width or 14
            )
    return wb


def main() -> None:
    results: dict[str, dict] = {}
    for tool_id, payload in SAMPLES.items():
        try:
            results[tool_id] = get_tool(tool_id)(**payload)
        except Exception as exc:  # noqa: BLE001
            results[tool_id] = {"error": str(exc)}

    out_dir = ROOT / "output"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"ket_qua_{stamp}.xlsx"
    template_path = ROOT / "excel" / "ket_qua_template.xlsx"

    wb = build_workbook(results)
    wb.save(out_path)

    # Also refresh committed template (sample run)
    template_path.parent.mkdir(exist_ok=True)
    wb.save(template_path)

    latest = out_dir / "ket_qua_latest.xlsx"
    wb.save(latest)
    print(f"Wrote {out_path}")
    print(f"Wrote {template_path}")
    print(f"Wrote {latest}")


if __name__ == "__main__":
    main()
